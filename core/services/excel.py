from copy import deepcopy
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from core.services.base import BaseDocumentService


class ExcelService(BaseDocumentService):
    def __init__(self):
        self.workbook = None

    def load(self, file) -> None:
        self.workbook = load_workbook(file)

    def update(self, params: dict) -> None:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.display in params:
                        value = params.get(cell.hyperlink.display)

                        if cell.hyperlink.location:
                            location_sheet, location_cell = (
                                cell.hyperlink.location.split("!")
                            )
                            location_sheet = location_sheet.replace("'", "")
                            location_sheet = self.workbook[location_sheet]
                            start_cell = location_sheet[location_cell]

                            if isinstance(
                                value, list
                            ):  # Если значение словаря - список из списков
                                for i, sublist in enumerate(
                                    value
                                ):  # Прокод по спискам списка
                                    for j, item in enumerate(
                                        sublist
                                    ):  # Проход по значениям из вложенного списка
                                        target_cell_col = start_cell.column + j
                                        target_cell_row = start_cell.row + i
                                        col_letter = get_column_letter(target_cell_col)
                                        self._write_value_to_cell(
                                            location_sheet,
                                            f"{col_letter}{target_cell_row}",
                                            item,
                                        )

                            else:
                                self._write_value_to_cell(sheet, location_cell, value)

    def to_json(self, sheet_name: str = None, range: str = None) -> dict:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")
        sheets = self.workbook.worksheets
        if sheet_name and sheet_name in self.workbook.sheetnames:
            sheets = [self.workbook[sheet_name]]

        data = {}

        for sheet in sheets:
            base_range = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            if range:
                base_range = range
            sheet_data = {}
            cells_range = sheet[base_range]
            for row in cells_range:
                for cell in row:
                    if cell.value:
                        cell_letter = get_column_letter(cell.column)
                        cell_address = f"{cell_letter}{cell.row}"
                        sheet_data[cell_address] = cell.value
            if sheet_data:
                data[sheet.title] = sheet_data

        return data

    def from_json(self, data: dict) -> None:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")

        for sheet_name, sheet_data in data.items():
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.create_sheet(sheet_name)

            for cell_address, value in sheet_data.items():
                try:
                    col_letter = "".join(
                        [char for char in cell_address if char.isalpha()]
                    )
                    row_number = int(
                        "".join([char for char in cell_address if char.isdigit()])
                    )
                    col_idx = column_index_from_string(col_letter)

                    if row_number < 1 or col_idx < 1:
                        raise ValueError(f"Некорректный адрес ячейки: {cell_address}.")

                    self._write_value_to_cell(sheet, cell_address, value)
                except Exception as e:
                    raise ValueError(f"Ошибка при обработке ячейки {cell_address}: {e}")

    def update_with_blocks(self, data: dict) -> None:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")

        original_workbook = deepcopy(self.workbook)

        blocks = data.get("blocks", [])
        newpage = data.get("newpage", True)
        merged_ranges = data.get("merged", [])
        format_data = data.get("format", {})

        sheet = self.workbook.active
        original_sheet = original_workbook.active

        if newpage:
            sheet_counter = 1
            for block in blocks:
                if sheet_counter > 1:
                    new_sheet = self.workbook.copy_worksheet(sheet)
                    new_sheet.title = f"{sheet.title}_{sheet_counter}"
                else:
                    new_sheet = sheet

                new_sheet.delete_rows(1, new_sheet.max_row)
                self._copy_sheet_content(original_sheet, new_sheet)

                self._process_block(new_sheet, block, merged_ranges, format_data)

                sheet_counter += 1
        else:
            for i, block in enumerate(blocks):
                self._copy_sheet_content(original_sheet, sheet)
                num_rows = original_sheet.max_row

                self._process_block(sheet, block, merged_ranges, format_data)

                if i < len(blocks) - 1:
                    sheet.insert_rows(1, num_rows + 1)

    def _process_block(self, sheet, block, merged_ranges, format_data):

        for cell_address, value in block.items():
            self._write_value_to_cell(sheet, cell_address, value)

        for merge_range in merged_ranges:
            self._merge_cells_with_data(sheet, merge_range)

        for cell_range, format_dict in format_data.items():
            if ":" in cell_range:
                for row in sheet[cell_range]:
                    for cell in row:
                        self._apply_formatting(cell, format_dict)
            else:
                self._apply_formatting(sheet[cell_range], format_dict)

    def _apply_formatting(self, cell, format_dict):
        font_args = {}

        if "textcolor" in format_dict:
            font_args["color"] = format_dict["textcolor"]

        if "fontname" in format_dict:
            font_args["name"] = format_dict["fontname"]

        if "fontsize" in format_dict:
            font_args["size"] = format_dict["fontsize"]

        if "bold" in format_dict:
            font_args["bold"] = format_dict["bold"]

        if "italic" in format_dict:
            font_args["italic"] = format_dict["italic"]

        if "underline" in format_dict:
            underline_value = format_dict["underline"]
            font_args["underline"] = "single" if underline_value else None

        if "strikethrough" in format_dict:
            font_args["strike"] = format_dict["strikethrough"]

        if font_args:
            cell.font = Font(**font_args)

        if "fillcolor" in format_dict:
            cell.fill = PatternFill(
                start_color=format_dict["fillcolor"],
                end_color=format_dict["fillcolor"],
                fill_type="solid",
            )

        alignment_args = {}

        if "align" in format_dict:
            alignment_args["horizontal"] = format_dict["align"]

        if "valign" in format_dict:
            alignment_args["vertical"] = format_dict["valign"]

        if alignment_args:
            cell.alignment = Alignment(**alignment_args)

    def _copy_sheet_content(self, source_sheet, target_sheet):
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                if isinstance(target_cell, MergedCell):
                    continue

                target_cell.value = cell.value

                if cell.has_style:
                    target_cell._style = cell._style
                if cell.hyperlink:
                    target_cell._hyperlink = cell.hyperlink
                if cell.comment:
                    target_cell.comment = cell.comment

        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

    def _write_value_to_cell(self, sheet, cell_address, value):
        target_cell = sheet[cell_address]
        if "\n" in str(value):
            target_cell.alignment = target_cell.alignment.copy(wrap_text=True)
        merged_range_to_recreate = None
        if isinstance(target_cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell_address in merged_range:
                    merged_range_to_recreate = str(merged_range)
                    sheet.unmerge_cells(merged_range_to_recreate)
                    break
        sheet[cell_address] = value
        if merged_range_to_recreate:
            sheet.merge_cells(merged_range_to_recreate)

    def _merge_cells_with_data(self, sheet, merge_range):
        min_col, min_row, max_col, max_row = range_boundaries(merge_range)

        merged_data = []
        for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                if cell.value:
                    merged_data.append(str(cell.value))

        combined_value = "\n".join(merged_data)
        cell_address = f"{get_column_letter(min_col)}{min_row}"
        self._write_value_to_cell(sheet, cell_address, combined_value)
        sheet.merge_cells(merge_range)

    def save_to_bytes(self) -> BytesIO:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def save_to_file(self, file_path: str) -> None:
        if self.workbook:
            self.workbook.save(file_path)
        else:
            raise ValueError("Excel файл не загружен.")
